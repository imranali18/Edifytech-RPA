from openpyxl import load_workbook


def t(row_num, col_num):
    return str(col_num) + str(row_num)


def down(row_num, col_num):
    return row_num + 1, col_num


def up(row_num, col_num):
    if row_num > 0:
        return row_num - 1, col_num
    return row_num, col_num


def right(row_num, col_num):
    return row_num, chr(ord(col_num) + 1)


def left(row_num, col_num):
    if col_num == "A":
        return row_num, col_num
    else:
        return row_num, chr(ord(col_num) - 1)

DSS_XLSX_FILE = r"..\..\Admin\Accommodations\CS1 Extra Time Accommodations.xlsx"
EXAM_SEATING_FILE = r"..\exam_seating.txt"
INIT_SHEET = "Sheet1"
INIT_ROW = 2
INIT_COL = "A"
FIELD_COUNT = 9 + 1
SEP_CNT = 2

# Load in the workbook
wb = load_workbook(DSS_XLSX_FILE)

ws = wb[INIT_SHEET]
row = INIT_ROW
col = INIT_COL

done = False
field_width = {"lastname": 0, "firstname": 0, "rcsid": 0, "rsection": 0, "building": 0, "room": 0, "section": 0,\
               "seat_row": 0, "seat_number": 0, "timeslot": 0}
student_list = []

while not done:
    if ws[t(row, col)].value is None:
        done = True
        continue

    lastname, firstname = [name.strip() for name in ws[t(row, col)].value.split(",")]
    lastname = lastname.replace(" ", "_")
    firstname = firstname.replace(" ", "_")
    field_width["lastname"] = max(field_width["lastname"], len(lastname))
    field_width["firstname"] = max(field_width["firstname"], len(firstname))

    row, col = right(row, col)
    rcsid = ws[t(row, col)].value.split("@")[0]
    field_width["rcsid"] = max(field_width["rcsid"], len(rcsid))

    row, col = right(row, col)
    rsection = str(ws[t(row, col)].value)
    field_width["rsection"] = max(field_width["rsection"], len(rsection))

    row, col = right(row, col)
    building = ws[t(row, col)].value.capitalize()
    field_width["building"] = max(field_width["building"], len(building))

    row, col = right(row, col)
    room = str(ws[t(row, col)].value)
    field_width["room"] = max(field_width["room"], len(room))

    row, col = right(row, col)
    section = str(ws[t(row, col)].value)
    field_width["section"] = max(field_width["section"], len(section))

    row, col = right(row, col)
    seat_row = str(ws[t(row, col)].value)
    field_width["seat_row"] = max(field_width["seat_row"], len(seat_row))

    row, col = right(row, col)
    seat_number = str(ws[t(row, col)].value)
    field_width["seat_number"] = max(field_width["seat_number"], len(seat_number))

    row, col = right(row, col)
    row, col = right(row, col)
    timeslot = ws[t(row, col)].value
    field_width["timeslot"] = max(field_width["timeslot"], len(timeslot))

    #print(lastname, firstname, rcsid, rsection, building, room, section, seat_row, seat_number, timeslot)
    student_list.append((lastname, firstname, rcsid, rsection, building, room, section, seat_row, \
                         seat_number, timeslot))

    row, col = down(row, col)
    col = INIT_COL

with open(EXAM_SEATING_FILE, "w") as seating_file:
    for student_data in student_list:
        lastname, firstname, rcsid, rsection, building, room, section, seat_row, seat_number, timeslot = student_data
        student_line = "{:{width1}s}{:{width2}s}{:{width3}s}{:{width4}s}{:{width5}s}" \
              "{:{width6}s}{:{width7}s}{:{width8}s}{:{width9}s}{:{width10}s}". \
              format(lastname, firstname, rcsid, rsection, building, room, section, seat_row, seat_number, timeslot, \
                     width1=field_width["lastname"] + SEP_CNT, width2=field_width["firstname"] + SEP_CNT, \
                     width3=field_width["rcsid"] + SEP_CNT, width4=field_width["rsection"] + SEP_CNT, \
                     width5=field_width["building"] + SEP_CNT, width6=field_width["room"] + SEP_CNT, \
                     width7=field_width["section"] + SEP_CNT, width8=field_width["seat_row"] + SEP_CNT, \
                     width9=field_width["seat_number"] + SEP_CNT, width10=field_width["timeslot"] + SEP_CNT)
        print(student_line)
        seating_file.write(student_line + "\n")

